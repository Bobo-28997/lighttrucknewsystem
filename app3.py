# =====================================
# Streamlit App: 人事用“项目提成 & 二次项目 & 平台工 & 独立架构 & 低价值”自动审核（V2 - 新规则版）
# - 移除 "权责发生" 和 "超期明细"
# - 新增 "完成二次交接时间" vs "出本流程时间" (日期)
# - 新增 "年化MIN" vs "XIRR_商_起租" (数值)
# - 新增 "年限" vs "租赁期限" (*12, 容错1个月)
# - 更改 "城市经理" -> 使用 "放款明细"
# - 移除 "产品" 检查
# =====================================

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import unicodedata, re
import time # (确保 time 被导入)

st.title("📊 模拟人事用薪资计算表自动审核系统-2 (新规则版)")

st.image("image/app2image.drawio.png")

# ========== 上传文件 ==========
# (修改点：文件数 5 -> 4, 移除 "超期明细")
uploaded_files = st.file_uploader(
    "请上传文件名中包含以下字段的文件：项目提成、放款明细、二次明细、产品台账。最后誊写，需检的表为项目提成表。",
    type="xlsx", accept_multiple_files=True
)

if not uploaded_files or len(uploaded_files) < 4:
    st.warning("⚠️ 请上传所有 4 个文件（项目提成、二次明细、放款明细、产品台账）")
    st.stop()
else:
    st.success("✅ 文件上传完成")

# ========== 工具函数 (不变) ==========
def find_file(files_list, keyword):
    for f in files_list:
        if keyword in f.name:
            return f
    raise FileNotFoundError(f"未找到包含关键词「{keyword}」的文件")

def normalize_colname(c):
    return str(c).strip().lower()

def find_col(df, keyword, exact=False):
    if df is None:
        return None
    key = keyword.strip().lower()
    for col in df.columns:
        cname = normalize_colname(col)
        if (exact and cname == key) or (not exact and key in cname):
            return col
    return None

def normalize_num(val):
    if pd.isna(val):
        return None
    s = str(val).replace(",", "").strip()
    if s in ["", "-", "nan"]:
        return None
    try:
        if "%" in s:
            s = s.replace("%", "")
            return float(s) / 100
        return float(s)
    except:
        return s

def normalize_text(val):
    if pd.isna(val):
        return ""
    s = str(val)
    s = re.sub(r'[\n\r\t ]+', '', s)
    s = s.replace('\u3000', '')
    s = ''.join(unicodedata.normalize('NFKC', ch) for ch in s)
    return s.lower().strip()

def detect_header_row(file, sheet_name):
    preview = pd.read_excel(file, sheet_name=sheet_name, nrows=2, header=None)
    first_row = preview.iloc[0]
    total_cells = len(first_row)
    empty_like = sum(
        (pd.isna(x) or str(x).startswith("Unnamed") or str(x).strip() == "")
        for x in first_row
    )
    empty_ratio = empty_like / total_cells if total_cells > 0 else 0
    return 1 if empty_ratio >= 0.7 else 0

def get_header_row(file, sheet_name):
    if any(k in sheet_name for k in ["起租", "二次"]):
        return 1
    return detect_header_row(file, sheet_name)

def normalize_contract_key(series: pd.Series) -> pd.Series:
    s = series.astype(str)
    s = s.str.replace(r"\.0$", "", regex=True) 
    s = s.str.strip()
    s = s.str.upper() 
    s = s.str.replace('－', '-', regex=False)
    return s

def prepare_one_ref_df(ref_df, ref_contract_col, required_cols, prefix):
    if ref_df is None:
        st.warning(f"⚠️ 参考文件 '{prefix}' 未加载 (df is None)。")
        return pd.DataFrame(columns=['__KEY__'])
        
    if ref_contract_col is None:
        st.warning(f"⚠️ 在 {prefix} 参考表中未找到'合同'列，跳过此数据源。")
        return pd.DataFrame(columns=['__KEY__'])

    cols_to_extract = []
    col_mapping = {} 

    for col_kw in required_cols:
        actual_col = find_col(ref_df, col_kw)
        
        if actual_col:
            cols_to_extract.append(actual_col)
            col_mapping[actual_col] = f"ref_{prefix}_{col_kw}"
        else:
            st.warning(f"⚠️ 在 {prefix} 参考表中未找到列 (关键字: '{col_kw}')")
            
    if not cols_to_extract:
        st.warning(f"⚠️ 在 {prefix} 参考表中未找到任何所需字段，跳过。")
        return pd.DataFrame(columns=['__KEY__'])

    cols_to_extract.append(ref_contract_col)
    cols_to_extract_unique = list(set(cols_to_extract))
    valid_cols = [col for col in cols_to_extract_unique if col in ref_df.columns]
    std_df = ref_df[valid_cols].copy()
    
    # --- (修改点：* 12 逻辑已被移除) ---
    # (原有的 if prefix == 'fk' and '租赁期限' in required_cols: ... 块已被删除)

    std_df['__KEY__'] = normalize_contract_key(std_df[ref_contract_col])
    std_df = std_df.rename(columns=col_mapping)
    final_cols = ['__KEY__'] + list(col_mapping.values())
    final_cols_in_df = [col for col in final_cols if col in std_df.columns]
    std_df = std_df[final_cols_in_df]
    std_df = std_df.drop_duplicates(subset=['__KEY__'], keep='first')
    return std_df

# =====================================
# 🛠️ (修改) 比较函数 (compare_series_vec)
# =====================================
def compare_series_vec(s_main, s_ref, compare_type='text', tolerance=0, multiplier=1):
    """
    (V3: 增加 'num_term' 类型)
    """
    merge_failed_mask = s_ref.isna()
    main_is_na = pd.isna(s_main) | (s_main.astype(str).str.strip().isin(["", "nan", "None"]))
    ref_is_na = pd.isna(s_ref) | (s_ref.astype(str).str.strip().isin(["", "nan", "None"]))
    both_are_na = main_is_na & ref_is_na
    
    errors = pd.Series(False, index=s_main.index)

    # 2. 日期比较
    if compare_type == 'date':
        d_main = pd.to_datetime(s_main, errors='coerce')
        d_ref = pd.to_datetime(s_ref, errors='coerce')
        
        valid_dates_mask = d_main.notna() & d_ref.notna()
        date_diff_mask = (d_main.dt.date != d_ref.dt.date)
        errors = valid_dates_mask & date_diff_mask
        
        one_is_date_one_is_not = (d_main.notna() & d_ref.isna() & ~ref_is_na) | \
                                 (d_main.isna() & ~main_is_na & d_ref.notna())
        errors |= one_is_date_one_is_not

    # 3. 数值比较 (修改点：增加 num_term)
    elif compare_type in ['num', 'num_term']:
        s_main_norm = s_main.apply(normalize_num)
        s_ref_norm = s_ref.apply(normalize_num)
        
        # (注意: 乘法 * 12 已移至 prepare_one_ref_df)
        
        is_num_main = s_main_norm.apply(lambda x: isinstance(x, (int, float)))
        is_num_ref = s_ref_norm.apply(lambda x: isinstance(x, (int, float)))
        both_are_num = is_num_main & is_num_ref

        if both_are_num.any():
            diff = (s_main_norm[both_are_num] - s_ref_norm[both_are_num]).abs()
            
            # --- (修改点：分离 'num_term' 逻辑) ---
            if compare_type == 'num_term':
                # "年限": 忽略小于1.0的差距 (>= 1.0 算错误)
                errors.loc[both_are_num] = (diff >= 1.0)
            else:
                # "年化MIN" / "收益率": 正常容错
                errors.loc[both_are_num] = (diff > (tolerance + 1e-6))
            # --- (修改结束) ---
                
        one_is_num_one_is_not = (is_num_main & ~is_num_ref & ~ref_is_na) | \
                                (~is_num_main & ~main_is_na & is_num_ref)
        errors |= one_is_num_one_is_not

    # 4. 文本比较
    else: # compare_type == 'text'
        s_main_norm_text = s_main.apply(normalize_text)
        s_ref_norm_text = s_ref.apply(normalize_text)
        errors = (s_main_norm_text != s_ref_norm_text)

    # 5. 最终错误逻辑
    final_errors = errors & ~both_are_na
    lookup_failure_mask = merge_failed_mask & ~main_is_na
    final_errors = final_errors & ~lookup_failure_mask
    
    return final_errors

# =====================================
# 🧮 审核函数 (不变)
# =====================================
def audit_sheet_vec(sheet_name, main_file, all_std_dfs, mapping_rules_vec):
    xls_main = pd.ExcelFile(main_file)
    
    header_offset = get_header_row(main_file, sheet_name)
    main_df = pd.read_excel(xls_main, sheet_name=sheet_name, header=header_offset)
    st.write(f"📘 审核中：{sheet_name}（header={header_offset}）")

    contract_col_main = find_col(main_df, "合同")
    if not contract_col_main:
        st.error(f"❌ {sheet_name} 中未找到“合同”列，已跳过。")
        return None, 0

    main_df['__ROW_IDX__'] = main_df.index
    main_df['__KEY__'] = normalize_contract_key(main_df[contract_col_main])

    merged_df = main_df.copy()
    for std_df in all_std_dfs.values():
        if not std_df.empty:
            merged_df = pd.merge(merged_df, std_df, on='__KEY__', how='left')

    total_errors = 0
    errors_locations = set()
    row_has_error = pd.Series(False, index=merged_df.index)

    progress = st.progress(0)
    status = st.empty()
    
    total_comparisons = len(mapping_rules_vec)
    current_comparison = 0

    for main_kw, comparisons in mapping_rules_vec.items():
        current_comparison += 1
        
        main_col = find_col(main_df, main_kw)
        if not main_col:
            continue 
        
        status.text(f"检查「{sheet_name}」: {main_kw}...")
        
        field_error_mask = pd.Series(False, index=merged_df.index)
        
        for (ref_col, compare_type, tol, mult) in comparisons:
            if ref_col not in merged_df.columns:
                continue 
            
            s_main = merged_df[main_col]
            s_ref = merged_df[ref_col]
            
            # (注意：这里的 mult 参数现在只用于旧的 'num' 类型, 'num_term' 的乘法已在预处理中完成)
            errors_mask = compare_series_vec(s_main, s_ref, compare_type, tol, mult)
            
            field_error_mask |= errors_mask
        
        if field_error_mask.any():
            total_errors += field_error_mask.sum()
            row_has_error |= field_error_mask
            
            bad_indices = merged_df[field_error_mask]['__ROW_IDX__']
            for idx in bad_indices:
                errors_locations.add((idx, main_col))
        
        progress.progress(current_comparison / total_comparisons)

    status.text(f"「{sheet_name}」比对完成，正在生成标注文件...")

    # 5. === 快速写入 Excel 并标注 ===
    wb = Workbook()
    ws = wb.active
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    original_cols_list = list(main_df.drop(columns=['__ROW_IDX__', '__KEY__']).columns)
    col_name_to_idx = {name: i + 1 for i, name in enumerate(original_cols_list)}
    
    if header_offset > 0:
        for _ in range(header_offset):
            ws.append([""] * len(original_cols_list))
            
    for r in dataframe_to_rows(main_df[original_cols_list], index=False, header=True):
        ws.append(r)

    for (row_idx, col_name) in errors_locations:
        if col_name in col_name_to_idx:
            excel_row = row_idx + 1 + header_offset + 1
            excel_col = col_name_to_idx[col_name]
            ws.cell(excel_row, excel_col).fill = red_fill
            
    if contract_col_main in col_name_to_idx:
        contract_col_excel_idx = col_name_to_idx[contract_col_main]
        error_row_indices = merged_df[row_has_error]['__ROW_IDX__']
        for row_idx in error_row_indices:
            excel_row = row_idx + 1 + header_offset + 1
            ws.cell(excel_row, contract_col_excel_idx).fill = yellow_fill

    # 6. 导出
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    st.download_button(
        label=f"📥 下载 {sheet_name} 审核标注版",
        data=output_stream,
        file_name=f"{sheet_name}_审核标注版.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_{sheet_name}"
    )

    # 7. (新) 导出仅含错误行的文件 (带标红)
    if row_has_error.any():
        try:
            df_errors_only = merged_df.loc[row_has_error, original_cols_list].copy()
            original_indices_with_error = merged_df.loc[row_has_error, '__ROW_IDX__']
            original_idx_to_new_excel_row = {
                original_idx: new_row_num 
                for new_row_num, original_idx in enumerate(original_indices_with_error, start=2)
            }
            wb_errors = Workbook()
            ws_errors = wb_errors.active
            for r in dataframe_to_rows(df_errors_only, index=False, header=True):
                ws_errors.append(r)
            for (original_row_idx, col_name) in errors_locations:
                if original_row_idx in original_idx_to_new_excel_row:
                    new_row = original_idx_to_new_excel_row[original_row_idx]
                    if col_name in col_name_to_idx:
                        new_col = col_name_to_idx[col_name]
                        ws_errors.cell(row=new_row, column=new_col).fill = red_fill
            
            output_errors_only = BytesIO()
            wb_errors.save(output_errors_only)
            output_errors_only.seek(0)
            
            st.download_button(
                label=f"📥 下载 {sheet_name} (仅含错误行, 带标红)",
                data=output_errors_only,
                file_name=f"{sheet_name}_仅错误行_标红.xlsx",
                key=f"download_{sheet_name}_errors_only"
            )
        except Exception as e:
            st.error(f"❌ 生成“仅错误行”文件时出错: {e}")
            
    st.success(f"✅ {sheet_name} 审核完成，共发现 {total_errors} 处错误")
    
    return main_df.drop(columns=['__ROW_IDX__', '__KEY__']), total_errors

# =====================================
# 🛠️ (修改) 文件读取 & 预处理 (V3 - 使用 "提成" sheet 并更新映射)
# =====================================
main_file = find_file(uploaded_files, "项目提成")
ec_file = find_file(uploaded_files, "二次明细")
fk_file = find_file(uploaded_files, "放款明细")
product_file = find_file(uploaded_files, "产品台账")

st.info("ℹ️ 正在读取并预处理参考文件...")

# 1. 加载所有参考 DF
ec_df = pd.read_excel(ec_file)
product_df = pd.read_excel(product_file)
fk_xls = pd.ExcelFile(fk_file)

# --- VVVV (【核心修改】加载 "提成" sheets) VVVV ---
commission_sheets = [s for s in fk_xls.sheet_names if "提成" in s]

if not commission_sheets:
    st.error("❌ 在 '放款明细' 文件中未找到任何包含 '提成' 的sheet！程序已停止。")
    st.stop()

st.info(f"ℹ️ 正在从 '放款明细' 加载 {len(commission_sheets)} 个 '提成' sheet...")

commission_df_list = [pd.read_excel(fk_xls, sheet_name=s) for s in commission_sheets]
fk_commission_df = pd.concat(commission_df_list, ignore_index=True)

# 将 fk_df 和 commission_df 都指向这个合并后的 DataFrame
fk_df = fk_commission_df         # <--- 用于字段验证
commission_df = fk_commission_df # <--- 用于漏填检查
# --- ^^^^ (修改结束) ^^^^ ---

# ---- 找到所有参考表的合同列 ----
contract_col_ec = find_col(ec_df, "合同")
contract_col_fk = find_col(fk_df, "合同")
contract_col_comm = find_col(commission_df, "合同")
contract_col_product = find_col(product_df, "合同")

# 2. (修改) 定义向量化映射规则
mapping_rules_vec = {
    "起租日期": [
        ("ref_ec_起租日_商", 'date', 0, 1)
    ],
    "租赁本金": [("ref_fk_租赁本金", 'num', 0, 1)],
    "收益率": [("ref_product_XIRR_商_起租", 'num', 0.005, 1)],
    
    # --- VVVV (【核心修改】映射到 "提报人员") VVVV ---
    "操作人": [("ref_fk_提报人员", 'text', 0, 1)],
    "客户经理": [("ref_fk_提报人员", 'text', 0, 1)],
    # --- ^^^^ (修改结束) ^^^^ ---
    
    "城市经理": [("ref_fk_城市经理", 'text', 0, 1)],
    "完成二次交接时间": [("ref_ec_出本流程时间", 'date', 0, 1)],
    "年化MIN": [("ref_product_XIRR_商_起租", 'num', 0.005, 1)],
    "年限": [("ref_fk_租赁期限", 'num_term', 0, 0)]
}

# 3. (修改) 预处理所有参考 DF
ec_cols = ["起租日_商", "出本流程时间"]

# --- VVVV (【核心修改】使用 "提报人员") VVVV ---
fk_cols = ["租赁本金", "提报人员", "城市经理", "租赁期限"] # <--- "客户经理" 已改为 "提报人员"
# --- ^^^^ (修改结束) ^^^^ ---

product_cols = ["起租日_商", "XIRR_商_起租"]

ec_std = prepare_one_ref_df(ec_df, contract_col_ec, ec_cols, "ec")
fk_std = prepare_one_ref_df(fk_df, contract_col_fk, fk_cols, "fk")
product_std = prepare_one_ref_df(product_df, contract_col_product, product_cols, "product")

all_std_dfs = {
    "ec": ec_std,
    "fk": fk_std,
    "product": product_std,
}

st.success("✅ 参考文件预处理完成。")

# =====================================
# 🛠️ (修改) 执行主流程
# =====================================
xls_main = pd.ExcelFile(main_file)
target_sheets = [
    s for s in xls_main.sheet_names
    if any(k in s for k in ["起租", "二次", "平台工", "独立架构", "低价值"]) # <--- 移除 "权责发生"
]

all_contracts_in_sheets = set()

if not target_sheets:
    st.warning("⚠️ 未找到目标 sheet。")
else:
    for sheet_name in target_sheets:
        df, _ = audit_sheet_vec(sheet_name, main_file, all_std_dfs, mapping_rules_vec)
        
        if df is not None:
            col = find_col(df, "合同")
            if col:
                normalized_contracts = normalize_contract_key(df[col].dropna())
                all_contracts_in_sheets.update(normalized_contracts)

# ======= (不变) 漏填检测 =======
if commission_df is not None and contract_col_comm:
    commission_contracts = set(normalize_contract_key(commission_df[contract_col_comm].dropna()))
    
    missing_contracts = sorted(list(commission_contracts - all_contracts_in_sheets))

    st.subheader("📋 合同漏填检测结果（基于提成sheet）")
    st.write(f"共 {len(missing_contracts)} 个合同在六张表中未出现。")

    if missing_contracts:
        wb_miss = Workbook()
        ws_miss = wb_miss.active
        ws_miss.cell(1, 1, "未出现在任一表中的合同号")
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, cno in enumerate(missing_contracts, start=2):
            ws_miss.cell(i, 1, cno).fill = yellow

        out_miss = BytesIO()
        wb_miss.save(out_miss)
        out_miss.seek(0)
        st.download_button(
            "📥 下载漏填合同列表",
            data=out_miss,
            file_name="漏填合同号列表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_missing_contracts" # (添加一个唯一的 key)
        )
    else:
        st.success("✅ 所有提成sheet合同号均已出现在六张表中，无漏填。")
